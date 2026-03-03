"""
Mapping 文件解析器
支持灵活适配不同表头的 Mapping 文档

Mapping 文件包含两个页签：
1. 表结构映射 - 来源表与目标表的关系
2. 字段级映射 - 来源字段到目标字段的映射

特性：
- 多级匹配策略：完全匹配 → 别名匹配 → 模糊匹配
- 支持外部 YAML 配置文件自定义列名映射
- 调试模式下输出详细的匹配信息
"""
import os
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, field, asdict
from openpyxl import load_workbook

from .header_config import HeaderConfigManager, ColumnMatch


# ============== 数据模型 ==============

@dataclass
class TableMapping:
    """表结构映射"""
    group_id: Optional[int] = None           # 分组 ID
    group: Optional[int] = None              # 分组
    source_schema: Optional[str] = None      # 来源 schema
    source_table: Optional[str] = None       # 来源表英文名
    source_table_cn: Optional[str] = None    # 来源表中文名
    source_alias: Optional[str] = None       # 来源表别名
    target_schema: Optional[str] = None      # 目标 schema
    target_table: Optional[str] = None       # 目标表英文名
    target_table_cn: Optional[str] = None    # 目标表中文名
    filter_condition: Optional[str] = None   # 取数条件
    join_condition: Optional[str] = None     # 关联条件

    def to_dict(self) -> Dict:
        return {k: v for k, v in asdict(self).items() if v is not None}


@dataclass
class FieldMapping:
    """字段级映射"""
    group_id: Optional[int] = None           # 分组 ID
    group: Optional[int] = None              # 分组
    source_schema: Optional[str] = None      # 来源 schema
    source_table: Optional[str] = None       # 来源表英文名
    source_alias: Optional[str] = None       # 来源表别名
    source_field: Optional[str] = None       # 来源字段英文名
    source_field_cn: Optional[str] = None    # 来源字段中文名
    source_field_type: Optional[str] = None  # 来源字段类型
    mapping_scene: Optional[str] = None      # 映射场景（直接复制/数据加工/赋值等）
    transform_rule: Optional[str] = None     # 具体加工规则
    target_schema: Optional[str] = None      # 目标 schema
    target_table: Optional[str] = None       # 目标表英文名
    target_field: Optional[str] = None       # 目标字段英文名
    target_field_cn: Optional[str] = None    # 目标字段中文名
    target_field_type: Optional[str] = None  # 目标字段类型

    def to_dict(self) -> Dict:
        return {k: v for k, v in asdict(self).items() if v is not None}


@dataclass
class ParseResult:
    """解析结果"""
    table_mappings: List[TableMapping] = field(default_factory=list)
    field_mappings: List[FieldMapping] = field(default_factory=list)
    source_tables: List[str] = field(default_factory=list)
    target_table: Optional[str] = None
    metadata: Dict = field(default_factory=dict)
    parse_report: Dict = field(default_factory=dict)  # 解析报告


# ============== Mapping 解析器 ==============

class MappingParser:
    """
    Mapping 文件解析器

    支持：
    1. 解析两个页签：表结构映射和字段级映射
    2. 灵活适配不同的表头列名（多级匹配策略）
    3. 支持外部配置文件自定义列名映射
    4. 调试模式下输出详细的匹配信息
    """

    def __init__(
        self, 
        config_path: Optional[str] = None,
        debug: bool = False
    ):
        """
        初始化解析器

        Args:
            config_path: 外部配置文件路径，可选
            debug: 是否输出调试信息
        """
        self.config = HeaderConfigManager(config_path)
        self.debug = debug
        self.table_mappings: List[TableMapping] = []
        self.field_mappings: List[FieldMapping] = []
        self._parse_report: Dict = {}

    def parse(self, file_path: str) -> Dict[str, Any]:
        """
        解析 Mapping 文件

        Args:
            file_path: Mapping 文件路径

        Returns:
            {
                'table_mappings': List[TableMapping],
                'field_mappings': List[FieldMapping],
                'source_tables': List[str],
                'target_table': str,
                'metadata': Dict,
                'parse_report': Dict
            }
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Mapping 文件不存在：{file_path}")

        wb = load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames

        result = ParseResult()
        result.metadata = {
            'file_path': file_path,
            'sheets': sheet_names,
        }

        # 解析表结构映射页签
        table_sheet = self.config.find_sheet(sheet_names, 'table')
        if table_sheet:
            ws = wb[table_sheet]
            result.table_mappings, table_report = self._parse_table_mapping(ws)
            result.metadata['table_sheet'] = table_sheet
            result.parse_report['table_mapping'] = table_report
        else:
            if self.debug:
                print(f"[Warning] 未找到表结构映射页签，可用页签: {sheet_names}")

        # 解析字段级映射页签
        field_sheet = self.config.find_sheet(sheet_names, 'field')
        if field_sheet:
            ws = wb[field_sheet]
            result.field_mappings, field_report = self._parse_field_mapping(ws)
            result.metadata['field_sheet'] = field_sheet
            result.parse_report['field_mapping'] = field_report
        else:
            if self.debug:
                print(f"[Warning] 未找到字段级映射页签，可用页签: {sheet_names}")

        # 提取来源表和目标表信息
        result.source_tables = self._extract_source_tables(result.table_mappings)
        result.target_table = self._extract_target_table(
            result.table_mappings, 
            result.field_mappings
        )

        # 保存到实例属性
        self.table_mappings = result.table_mappings
        self.field_mappings = result.field_mappings

        if self.debug:
            self._print_parse_report(result)

        return result.to_dict() if hasattr(result, 'to_dict') else self._result_to_dict(result)

    def _result_to_dict(self, result: ParseResult) -> Dict[str, Any]:
        """将 ParseResult 转换为字典"""
        return {
            'table_mappings': [m.to_dict() for m in result.table_mappings],
            'field_mappings': [m.to_dict() for m in result.field_mappings],
            'source_tables': result.source_tables,
            'target_table': result.target_table,
            'metadata': result.metadata,
            'parse_report': result.parse_report,
        }

    def _parse_table_mapping(self, ws) -> tuple:
        """
        解析表结构映射页签

        Args:
            ws: openpyxl worksheet 对象

        Returns:
            (List[TableMapping], 解析报告)
        """
        mappings = []
        report = {'headers': [], 'column_matches': {}}

        # 读取表头
        headers = [cell.value for cell in ws[1]]
        report['headers'] = headers

        # 使用 HeaderConfigManager 查找列索引
        col_matches = self.config.find_all_columns(headers, 'table')
        report['column_matches'] = {
            name: {
                'actual_name': match.actual_name,
                'index': match.index,
                'match_type': match.match_type
            }
            for name, match in col_matches.items()
        }

        if self.debug:
            print("\n" + self.config.get_match_report(headers, 'table'))

        # 构建列索引映射
        col_indices = {name: match.index for name, match in col_matches.items()}

        # 读取数据行
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            # 跳过空行
            if self._is_empty_row(row):
                continue

            mapping = TableMapping(
                group_id=self._get_int_value(row, col_indices.get('group_id')),
                group=self._get_int_value(row, col_indices.get('group')),
                source_schema=self._get_cell_value(row, col_indices.get('source_schema')),
                source_table=self._get_cell_value(row, col_indices.get('source_table')),
                source_table_cn=self._get_cell_value(row, col_indices.get('source_table_cn')),
                source_alias=self._get_cell_value(row, col_indices.get('source_alias')),
                target_schema=self._get_cell_value(row, col_indices.get('target_schema')),
                target_table=self._get_cell_value(row, col_indices.get('target_table')),
                target_table_cn=self._get_cell_value(row, col_indices.get('target_table_cn')),
                filter_condition=self._get_cell_value(row, col_indices.get('filter_condition')),
                join_condition=self._get_cell_value(row, col_indices.get('join_condition')),
            )

            # 只添加有实际内容的映射
            if mapping.source_table or mapping.target_table:
                mappings.append(mapping)

        return mappings, report

    def _parse_field_mapping(self, ws) -> tuple:
        """
        解析字段级映射页签

        Args:
            ws: openpyxl worksheet 对象

        Returns:
            (List[FieldMapping], 解析报告)
        """
        mappings = []
        report = {'headers': [], 'column_matches': {}}

        # 读取表头
        headers = [cell.value for cell in ws[1]]
        report['headers'] = headers

        # 使用 HeaderConfigManager 查找列索引
        col_matches = self.config.find_all_columns(headers, 'field')
        report['column_matches'] = {
            name: {
                'actual_name': match.actual_name,
                'index': match.index,
                'match_type': match.match_type
            }
            for name, match in col_matches.items()
        }

        if self.debug:
            print("\n" + self.config.get_match_report(headers, 'field'))

        # 构建列索引映射
        col_indices = {name: match.index for name, match in col_matches.items()}

        # 读取数据行
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            # 跳过空行
            if self._is_empty_row(row):
                continue

            mapping = FieldMapping(
                group_id=self._get_int_value(row, col_indices.get('group_id')),
                group=self._get_int_value(row, col_indices.get('group')),
                source_schema=self._get_cell_value(row, col_indices.get('source_schema')),
                source_table=self._get_cell_value(row, col_indices.get('source_table')),
                source_alias=self._get_cell_value(row, col_indices.get('source_alias')),
                source_field=self._get_cell_value(row, col_indices.get('source_field')),
                source_field_cn=self._get_cell_value(row, col_indices.get('source_field_cn')),
                source_field_type=self._get_cell_value(row, col_indices.get('source_field_type')),
                mapping_scene=self._get_cell_value(row, col_indices.get('mapping_scene')),
                transform_rule=self._get_cell_value(row, col_indices.get('transform_rule')),
                target_schema=self._get_cell_value(row, col_indices.get('target_schema')),
                target_table=self._get_cell_value(row, col_indices.get('target_table')),
                target_field=self._get_cell_value(row, col_indices.get('target_field')),
                target_field_cn=self._get_cell_value(row, col_indices.get('target_field_cn')),
                target_field_type=self._get_cell_value(row, col_indices.get('target_field_type')),
            )

            # 只添加有实际内容的映射
            if mapping.target_field or mapping.source_field:
                mappings.append(mapping)

        return mappings, report

    def _get_cell_value(self, row, col_index: Optional[int]) -> Optional[Any]:
        """获取单元格值，处理空值"""
        if col_index is None or col_index < 0:
            return None
        if col_index >= len(row):
            return None
        value = row[col_index].value
        if value is None:
            return None
        # 转换为字符串并去除首尾空格
        if isinstance(value, str):
            return value.strip() if value.strip() else None
        return value

    def _get_int_value(self, row, col_index: Optional[int]) -> Optional[int]:
        """获取单元格整数值"""
        value = self._get_cell_value(row, col_index)
        if value is None:
            return None
        try:
            return int(value)
        except (ValueError, TypeError):
            return None

    def _is_empty_row(self, row) -> bool:
        """判断是否为空行"""
        return all(cell.value is None for cell in row)

    def _extract_source_tables(self, table_mappings: List[TableMapping]) -> List[str]:
        """从表映射中提取来源表列表"""
        tables = set()
        for mapping in table_mappings:
            if mapping.source_table:
                # 添加 schema 前缀
                full_name = f"{mapping.source_schema}.{mapping.source_table}" if mapping.source_schema else mapping.source_table
                tables.add(full_name)
        return sorted(list(tables))

    def _extract_target_table(
        self, 
        table_mappings: List[TableMapping], 
        field_mappings: List[FieldMapping]
    ) -> Optional[str]:
        """提取目标表"""
        # 优先从表映射中获取
        for mapping in table_mappings:
            if mapping.target_table:
                if mapping.target_schema:
                    return f"{mapping.target_schema}.{mapping.target_table}"
                return mapping.target_table

        # 从字段映射中获取
        for mapping in field_mappings:
            if mapping.target_table:
                if mapping.target_schema:
                    return f"{mapping.target_schema}.{mapping.target_table}"
                return mapping.target_table

        return None

    def _print_parse_report(self, result: ParseResult) -> None:
        """打印解析报告"""
        print("\n" + "=" * 60)
        print("Mapping 解析报告")
        print("=" * 60)
        print(f"文件: {result.metadata.get('file_path', 'N/A')}")
        print(f"表结构映射页签: {result.metadata.get('table_sheet', 'N/A')}")
        print(f"字段级映射页签: {result.metadata.get('field_sheet', 'N/A')}")
        print(f"表映射记录数: {len(result.table_mappings)}")
        print(f"字段映射记录数: {len(result.field_mappings)}")
        print(f"来源表数量: {len(result.source_tables)}")
        print(f"目标表: {result.target_table or 'N/A'}")
        print("=" * 60)

    # ============== 查询方法 ==============

    def get_table_mappings(self) -> List[TableMapping]:
        """获取表结构映射"""
        return self.table_mappings

    def get_field_mappings(self) -> List[FieldMapping]:
        """获取字段级映射"""
        return self.field_mappings

    def get_mapping_by_target_field(self, field_name: str) -> Optional[FieldMapping]:
        """根据目标字段名获取映射"""
        for mapping in self.field_mappings:
            if mapping.target_field == field_name:
                return mapping
        return None

    def get_mappings_by_source_table(self, table_name: str) -> List[FieldMapping]:
        """根据来源表名获取相关字段映射"""
        return [m for m in self.field_mappings if m.source_table == table_name]

    def get_mappings_by_target_table(self, table_name: str) -> List[FieldMapping]:
        """根据目标表名获取相关字段映射"""
        return [m for m in self.field_mappings if m.target_table == table_name]

    def get_mappings_by_scene(self, scene: str) -> List[FieldMapping]:
        """根据映射场景获取字段映射"""
        return [m for m in self.field_mappings if m.mapping_scene == scene]

    def get_direct_copy_fields(self) -> List[FieldMapping]:
        """获取直接复制的字段映射"""
        return [m for m in self.field_mappings 
                if m.mapping_scene in ['直接复制', '直接映射', 'DIRECT', 'copy']]

    def get_transform_fields(self) -> List[FieldMapping]:
        """获取需要转换的字段映射"""
        return [m for m in self.field_mappings 
                if m.mapping_scene in ['数据加工', '转换', 'CALC', 'transform', 'TRANSFORM']]

    def get_constant_fields(self) -> List[FieldMapping]:
        """获取常量赋值的字段映射"""
        return [m for m in self.field_mappings 
                if m.mapping_scene in ['赋值', '常量', 'CONST', 'constant']]

    def get_primary_key_fields(self) -> List[FieldMapping]:
        """获取主键字段（根据字段名推断）"""
        pk_keywords = ['_ID', 'ID_', '_KEY', 'KEY_', '_NO', 'NO_', 'CODE']
        return [m for m in self.field_mappings 
                if any(kw in m.target_field.upper() for kw in pk_keywords) if m.target_field]

    def to_dict(self) -> Dict:
        """导出为字典"""
        return {
            'table_mappings': [m.to_dict() for m in self.table_mappings],
            'field_mappings': [m.to_dict() for m in self.field_mappings],
            'source_tables': self._extract_source_tables(self.table_mappings),
            'target_table': self._extract_target_table(self.table_mappings, self.field_mappings),
        }

    def to_json(self) -> str:
        """导出为 JSON 字符串"""
        import json
        return json.dumps(self.to_dict(), ensure_ascii=False, indent=2)

    # ============== 对接 MappingProcessor ==============

    def _infer_rule_type(self, mapping_scene: Optional[str], transform_rule: Optional[str] = None) -> str:
        """
        从映射场景推断规则类型

        Args:
            mapping_scene: 映射场景（直接复制/数据加工/赋值等）
            transform_rule: 加工规则（用于辅助判断）

        Returns:
            规则类型: DIRECT/CALC/AGG/JOIN/CASE/CONST/FUNC/SUBQ
        """
        # 先从加工规则推断（优先级更高，因为更精确）
        if transform_rule:
            expr_type = self._infer_rule_type_from_expression(transform_rule)
            if expr_type != 'DIRECT':  # 如果检测到特殊类型，优先使用
                return expr_type

        if not mapping_scene:
            return 'DIRECT'

        # 映射场景到规则类型的映射
        scene_mapping = {
            # 直接复制类
            '直接复制': 'DIRECT',
            '直接映射': 'DIRECT',
            '直取': 'DIRECT',
            '复制': 'DIRECT',
            'copy': 'DIRECT',
            'DIRECT': 'DIRECT',

            # 计算类
            '数据加工': 'CALC',
            '计算': 'CALC',
            '加工': 'CALC',
            '转换': 'CALC',
            'CALC': 'CALC',
            'transform': 'CALC',

            # 聚合类
            '聚合': 'AGG',
            '汇总': 'AGG',
            'AGG': 'AGG',

            # 关联类
            '关联': 'JOIN',
            '多表关联': 'JOIN',
            'JOIN': 'JOIN',

            # 条件类
            '条件': 'CASE',
            '条件映射': 'CASE',
            'CASE': 'CASE',

            # 常量类
            '赋值': 'CONST',
            '常量': 'CONST',
            '固定值': 'CONST',
            'CONST': 'CONST',
            'constant': 'CONST',

            # 函数类
            '函数': 'FUNC',
            'FUNC': 'FUNC',

            # 子查询类
            '子查询': 'SUBQ',
            'SUBQ': 'SUBQ',
        }

        # 标准化映射场景
        normalized_scene = mapping_scene.strip().upper() if mapping_scene else ''

        # 精确匹配
        if mapping_scene in scene_mapping:
            return scene_mapping[mapping_scene]

        # 大小写不敏感匹配
        if normalized_scene in scene_mapping:
            return scene_mapping[normalized_scene]

        # 部分匹配
        for key, rule_type in scene_mapping.items():
            if key.upper() in normalized_scene or normalized_scene in key.upper():
                return rule_type

        return 'DIRECT'

    def _infer_rule_type_from_expression(self, expression: str) -> str:
        """
        从 SQL 表达式推断规则类型

        Args:
            expression: SQL 表达式

        Returns:
            规则类型
        """
        if not expression:
            return 'DIRECT'

        expr_upper = expression.upper()

        # 检查子查询（多个 SELECT）
        if expr_upper.count('SELECT') > 1:
            return 'SUBQ'

        # 检查聚合函数
        agg_funcs = ['SUM(', 'COUNT(', 'AVG(', 'MAX(', 'MIN(', 'SUM (', 'COUNT (', 'AVG (']
        if any(func in expr_upper for func in agg_funcs):
            return 'AGG'

        # 检查 CASE WHEN
        if 'CASE' in expr_upper and 'WHEN' in expr_upper:
            return 'CASE'

        # 检查 JOIN
        if 'JOIN' in expr_upper:
            return 'JOIN'

        # 检查常用函数（包括空值处理函数）
        func_keywords = ['NVL', 'COALESCE', 'NULLIF', 'CAST', 'CONCAT', 'SUBSTR',
                         'TRIM', 'UPPER', 'LOWER', 'DATE_FORMAT', 'TO_DATE', 
                         'IFNULL', 'DECODE', 'NVL2']
        if any(func in expr_upper for func in func_keywords):
            return 'FUNC'

        # 检查计算符
        calc_operators = ['+', '-', '*', '/', '||']
        if any(op in expression for op in calc_operators):
            return 'CALC'

        return 'DIRECT'

    def to_processor_format(self) -> List[Dict[str, Any]]:
        """
        转换为 MappingProcessor 所需格式

        Returns:
            List[Dict] 符合 MappingProcessor.parse_from_dict() 要求的数据格式
        """
        result = []

        for m in self.field_mappings:
            # 推断规则类型
            rule_type = self._infer_rule_type(m.mapping_scene, m.transform_rule)

            # 构建转换规则（transform_rule 或 source_field）
            transformation = m.transform_rule
            if not transformation:
                if m.source_field:
                    transformation = m.source_field
                elif m.mapping_scene in ['赋值', '常量', 'CONST', 'constant']:
                    transformation = m.transform_rule or ''

            item = {
                'source_table': m.source_table or '',
                'source_field': m.source_field or '',
                'target_table': m.target_table or '',
                'target_field': m.target_field or '',
                'transformation_rule': transformation or '',
                'rule_type': rule_type,
                # 附加信息（供后续使用）
                'source_schema': m.source_schema,
                'target_schema': m.target_schema,
                'source_field_type': m.source_field_type,
                'target_field_type': m.target_field_type,
                'mapping_scene': m.mapping_scene,
            }
            result.append(item)

        return result

    def get_processor(self) -> 'MappingProcessor':
        """
        获取配置好的 MappingProcessor 实例

        这是便捷方法，一步完成：解析结果 -> 分层处理

        Returns:
            MappingProcessor 实例，已加载当前解析的 Mapping 数据

        Example:
            >>> parser = MappingParser()
            >>> parser.parse('mapping.xlsx')
            >>> processor = parser.get_processor()
            >>> # 按需查询三层信息
            >>> level1 = processor.get_level1_metadata({'is_primary_key': True})
        """
        from .mapping_processor import MappingProcessor

        processor = MappingProcessor()
        processor.parse_from_dict(self.to_processor_format())

        return processor


# ============== 便捷函数 ==============

def parse_mapping_file(
    file_path: str, 
    config_path: Optional[str] = None,
    debug: bool = False
) -> Dict[str, Any]:
    """
    解析 Mapping 文件的便捷函数
    
    Args:
        file_path: Mapping 文件路径
        config_path: 外部配置文件路径，可选
        debug: 是否输出调试信息
        
    Returns:
        解析结果字典
    """
    parser = MappingParser(config_path=config_path, debug=debug)
    return parser.parse(file_path)


def get_default_config_path() -> Optional[str]:
    """获取默认配置文件路径"""
    # 配置文件相对于当前文件的位置
    config_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
        'config', 
        'mapping_headers.yaml'
    )
    return config_path if os.path.exists(config_path) else None