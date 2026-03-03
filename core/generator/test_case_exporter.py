"""
测试用例导出器
支持导出为 Excel 格式
"""
import os
from typing import List, Dict
from datetime import datetime
from ..models import TestCase, TestCaseSuite


class TestCaseExporter:
    """测试用例导出器"""

    # Excel 列定义
    EXCEL_COLUMNS = [
        ("用例编号", "case_id"),
        ("用例名称", "case_name"),
        ("分类", "category"),
        ("测试场景", "scene"),
        ("优先级", "priority"),
        ("测试要点", "description"),
        ("涉及表", "tables"),
        ("前置条件", "pre_condition"),
        ("测试步骤 (SQL)", "test_steps"),
        ("预期结果", "expected_result"),
        ("后置条件", "post_condition"),
        ("状态", "status"),
        ("实际结果", "actual_result"),
        ("执行人", "executed_by"),
        ("执行时间", "executed_at"),
    ]

    def __init__(self):
        pass

    def export_to_excel(self, suite: TestCaseSuite, output_path: str) -> None:
        """
        导出测试用例集到 Excel 文件

        Args:
            suite: 测试用例集
            output_path: 输出文件路径
        """
        try:
            import pandas as pd
        except ImportError:
            raise ImportError("需要安装 pandas: pip install pandas openpyxl")

        # 确保输出目录存在
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)

        # 准备数据
        data = suite.to_excel_data()

        # 创建 DataFrame
        df = pd.DataFrame(data)

        # 重命名列
        column_mapping = {v: k for k, v in self.EXCEL_COLUMNS}
        df = df.rename(columns=column_mapping)

        # 确保列顺序
        df = df[[k for k, _ in self.EXCEL_COLUMNS]]

        # 导出到 Excel
        df.to_excel(output_path, index=False, sheet_name="测试用例")

        # 添加格式 (可选)
        self._format_excel(output_path, suite)

    def _format_excel(self, file_path: str, suite: TestCaseSuite) -> None:
        """
        格式化 Excel (添加表头信息等)

        Args:
            file_path: Excel 文件路径
            suite: 测试用例集
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")

        wb = load_workbook(file_path)
        ws = wb.active

        # 插入表头信息
        ws.insert_rows(0, 3)

        # 标题
        ws.merge_cells('A1:P1')
        ws['A1'] = f"数仓测试用例集 - {suite.target_table}"
        ws['A1'].font = ws['A1'].font.copy(bold=True, size=16)

        # 元信息
        ws['A2'] = f"设计版本：{suite.design_version}"
        ws['D2'] = f"生成时间：{suite.created_at.strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"用例总数：{len(suite.cases)}"
        ws['D3'] = f"高优先级：{len(suite.get_cases_by_priority('high'))}"

        # 自动调整列宽
        for col_idx in range(1, len(self.EXCEL_COLUMNS) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(file_path)

    def export_to_csv(self, suite: TestCaseSuite, output_path: str) -> None:
        """
        导出测试用例集到 CSV 文件

        Args:
            suite: 测试用例集
            output_path: 输出文件路径
        """
        try:
            import pandas as pd
        except ImportError:
            raise ImportError("需要安装 pandas: pip install pandas")

        # 确保输出目录存在
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)

        data = suite.to_excel_data()
        df = pd.DataFrame(data)

        # 重命名列
        column_mapping = {v: k for k, v in self.EXCEL_COLUMNS}
        df = df.rename(columns=column_mapping)

        # 导出到 CSV
        df.to_csv(output_path, index=False, encoding='utf-8-sig')

    def export_to_dict(self, suite: TestCaseSuite) -> List[Dict]:
        """
        导出为字典列表

        Args:
            suite: 测试用例集

        Returns:
            字典列表
        """
        return suite.to_excel_data()

    @staticmethod
    def create_test_case_from_dict(data: Dict) -> TestCase:
        """
        从字典创建测试用例

        Args:
            data: 用例数据字典

        Returns:
            TestCase 对象
        """
        tables = data.get("tables", [])
        if isinstance(tables, str):
            tables = tables.split(";")

        return TestCase(
            case_id=data.get("case_id", ""),
            case_name=data.get("case_name", ""),
            category=data.get("category", ""),
            scene=data.get("scene", ""),
            priority=data.get("priority", ""),
            description=data.get("description", ""),
            tables=tables,
            pre_condition=data.get("pre_condition", ""),
            test_steps=data.get("test_steps", ""),
            expected_result=data.get("expected_result", ""),
            post_condition=data.get("post_condition", ""),
        )