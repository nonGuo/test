import os
from openpyxl import load_workbook

os.chdir('d:/Projects/ai/test')

# 正确的文件名：mapping样例.xlsx（没有空格）
file_path = 'mapping样例.xlsx'
print(f'File exists: {os.path.exists(file_path)}')

wb = load_workbook(file_path, data_only=True)
print('Sheet names:', wb.sheetnames)

for sheet_name in wb.sheetnames:
    print(f'\n=== {sheet_name} ===')
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    print('Headers:', headers)
    print('Data (first 10 rows):')
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=11), 1):
        print(f'  Row {i}:', [cell.value for cell in row])
